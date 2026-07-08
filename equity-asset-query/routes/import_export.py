"""
导入导出路由模块
包含所有台账、天九持股主体、股权历史表的导入导出功能路由。
"""
import io
import json
import os
from datetime import datetime

import openpyxl
import requests
from flask import Blueprint, request, jsonify, send_file

import config
from shared import _current_env, _add_log, ensure_valid_token
from utils import _convert_inline_to_shared_strings, _parse_fail_file

import_export_bp = Blueprint('import_export', __name__)


@import_export_bp.route("/api/import/upload", methods=["POST"])
def import_upload():
    """上传导入文件到远程服务"""
    # 仅测试环境支持
    if _current_env["value"] != "test":
        return jsonify({"ok": False, "msg": "导入功能仅测试环境支持"})

    token = ensure_valid_token()
    if not token:
        return jsonify({"ok": False, "msg": "未登录或 Token 失效"})

    if "file" not in request.files:
        return jsonify({"ok": False, "msg": "未选择文件"})

    file = request.files["file"]
    if not file.filename:
        return jsonify({"ok": False, "msg": "文件名为空"})

    tenant = config.get_tenant(_current_env["value"])
    base_url = config.ENVIRONMENTS[_current_env["value"]]["base_url"]
    upload_url = f"{base_url}/tojoy-form-engine/data/uploadImportExcelFile"

    headers = {"Tj-Auth": token, "Tojoy-Tenant": tenant}
    files_data = {"files": (file.filename, file.stream,
                            file.content_type or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    form_data = {
        "appId": config.APP_ID,
        "formId": "5fe04116f79b468a87edc1e5e6ae5797",
        "importMode": "1",
        "saveMode": "1",
    }

    _add_log("info", "导入-上传", f"文件={file.filename}")
    try:
        resp = requests.post(upload_url, headers=headers, files=files_data, data=form_data, timeout=60)
        data = resp.json()
        if data.get("code") == 200:
            resp_data = data.get("data") or {}
            batch_no = resp_data.get("batchNo", "") if isinstance(resp_data, dict) else str(resp_data)
            _add_log("info", "导入-上传", f"成功: batchNo={batch_no}")
            return jsonify({"ok": True, "batchNo": batch_no, "data": resp_data, "msg": "文件上传成功"})
        else:
            _add_log("warn", "导入-上传", f"失败: {data.get('msg', '未知错误')}")
            return jsonify({"ok": False, "msg": f"上传失败：{data.get('msg', '未知错误')}"})
    except Exception as e:
        _add_log("error", "导入-上传", str(e))
        return jsonify({"ok": False, "msg": f"上传异常：{e}"})


@import_export_bp.route("/api/import/execute", methods=["POST"])
def import_execute():
    """执行台账导入"""
    # 仅测试环境支持
    if _current_env["value"] != "test":
        return jsonify({"ok": False, "msg": "导入功能仅测试环境支持"})

    token = ensure_valid_token()
    if not token:
        return jsonify({"ok": False, "msg": "未登录或 Token 失效"})

    body = request.get_json(force=True) or {}
    batch_no = body.get("batchNo", "")
    if not batch_no:
        return jsonify({"ok": False, "msg": "缺少 batchNo"})

    tenant = config.get_tenant(_current_env["value"])
    base_url = config.ENVIRONMENTS[_current_env["value"]]["base_url"]
    import_url = f"{base_url}/tojoy-pmis-service/ext/pmis/stockRight/importExcel"

    headers = {"Content-Type": "application/json", "Tj-Auth": token, "Tojoy-Tenant": tenant}
    payload = {
        "formId": "5fe04116f79b468a87edc1e5e6ae5797",
        "isStartTrigger": True,
        "appId": config.APP_ID,
        "isStockRightMain": True,
        "batchNo": batch_no,
        "importMode": "1",
        "saveMode": "1",
    }

    _add_log("info", "导入-执行", f"batchNo={batch_no}")
    try:
        resp = requests.post(import_url, json=payload, headers=headers, timeout=120)
        data = resp.json()
        resp_data = data.get("data") or {}

        # 提取统计信息（兼容不同字段名）
        total_num = resp_data.get("totalNum") or resp_data.get("total_num") or 0
        suc_num = resp_data.get("sucNum") or resp_data.get("suc_num") or resp_data.get("checkSucNum") or 0
        fail_num = resp_data.get("failNum") or resp_data.get("fail_num") or resp_data.get("checkFailNum") or 0
        # 失败文件URL（兼容不同字段名）
        fail_file_url = (resp_data.get("importFailFileUrl") or resp_data.get("failFileUrl")
                         or resp_data.get("fail_file_url") or "")
        import_result = resp_data.get("importResult")

        # 判断是否成功
        is_success = data.get("code") == 200 and (resp_data.get("success") is not False) and fail_num == 0 and suc_num > 0

        if is_success:
            _add_log("info", "导入-执行", f"成功: 共{total_num}行，入库{suc_num}行")
            return jsonify({"ok": True, "data": resp_data, "msg": f"导入成功：共{total_num}行，全部入库"})

        # 失败处理
        summary = f"共{total_num}行，通过{suc_num}行，失败{fail_num}行"
        if not total_num and not fail_num:
            summary = data.get("msg", "未知错误")

        # 尝试下载失败详情文件，解析失败原因
        fail_reasons = []
        if fail_file_url:
            fail_reasons = _parse_fail_file(fail_file_url)

        _add_log("warn", "导入-执行", f"失败: {summary}")
        return jsonify({
            "ok": False,
            "msg": f"导入失败：{summary}",
            "detail": {
                "totalNum": total_num,
                "sucNum": suc_num,
                "failNum": fail_num,
                "failFileUrl": fail_file_url,
                "failReasons": fail_reasons,
            },
        })
    except Exception as e:
        _add_log("error", "导入-执行", str(e))
        return jsonify({"ok": False, "msg": f"导入异常：{e}"})


@import_export_bp.route("/api/export/ledger-import-format", methods=["POST"])
def export_ledger_import_format():
    """按导入模板格式导出台账表当前页数据为 xlsx（支持生产和测试环境）
    使用 股权变更台账辅助表.xlsx 模板文件，保留其原始格式，从第3行开始填充数据。
    典型场景：从生产环境导出，导入到测试环境。
    """
    from flask import current_app

    # 获取当前台账表数据（从前端传过来）
    body = request.get_json(force=True) or {}
    records = body.get("records", [])
    if not records:
        return jsonify({"ok": False, "msg": "无数据可导出"})

    try:
        template_path = os.path.join(current_app.static_folder, "股权变更台账辅助表.xlsx")
        if not os.path.exists(template_path):
            _add_log("error", "导出", f"模板文件不存在: {template_path}")
            return jsonify({"ok": False, "msg": "导出模板文件不存在"})

        _add_log("info", "导出", f"环境={_current_env['value']}, 数据{len(records)}条")

        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        # 按台账表序号（number_8e6771）正序排列
        def sort_key(r):
            v = r.get("number_8e6771")
            if v is None:
                return float("inf")
            try:
                return float(v)
            except (ValueError, TypeError):
                return float("inf")
        records_sorted = sorted(records, key=sort_key)

        # 按导入模板列映射填充数据，从第3行开始（第1-2行为模板表头，含合并单元格）
        template_cols = config.LEDGER_IMPORT_TEMPLATE_COLS
        # 导出时清空的字段（编码类字段会干扰业务数据）
        SKIP_FIELDS = {"project_company_main_code", "input_15fbdf",
                       "agreement_subject_company_code", "ultimately_subject_company_code"}

        for row_idx, record in enumerate(records_sorted, start=3):
            for col_idx, (_, field, ftype) in enumerate(template_cols, start=1):
                if not field:
                    continue
                # 清空干扰业务数据的编码字段
                if field in SKIP_FIELDS:
                    ws.cell(row=row_idx, column=col_idx, value="")
                    continue
                val = record.get(field)
                # label 类型取 .label
                if ftype == "label" and isinstance(val, dict):
                    val = val.get("label", "")
                # 日期类型转为字符串
                if ftype == "date" and val:
                    s = str(val)
                    if s.isdigit() and len(s) >= 10:
                        ts = int(s)
                        if ts > 9999999999:
                            ts = ts / 1000
                        d = datetime.fromtimestamp(ts)
                        val = d.strftime("%Y-%m-%d")
                    elif len(s) >= 10 and "-" in s:
                        val = s[:10]
                if val is None:
                    val = ""
                # 确保值类型兼容 openpyxl（dict/list 转为字符串）
                if isinstance(val, (dict, list)):
                    val = json.dumps(val, ensure_ascii=False)
                ws.cell(row=row_idx, column=col_idx, value=val)

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        wb.close()

        # 后处理：将 openpyxl 生成的 inlineStr 格式转换为 sharedStrings 格式
        output = _convert_inline_to_shared_strings(output)

        # 文件名：导出台账-年月日时分
        now_str = datetime.now().strftime("%Y%m%d%H%M")
        download_name = f"导出台账-{now_str}.xlsx"

        _add_log("info", "导出", f"成功: {download_name}, {len(records_sorted)}条数据")
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=download_name,
        )
    except Exception as e:
        _add_log("error", "导出", str(e))
        return jsonify({"ok": False, "msg": f"导出失败：{e}"})


@import_export_bp.route("/api/export/holding-subject-format", methods=["POST"])
def export_holding_subject_format():
    """按导入模板格式导出天九持股主体表数据为 xlsx"""
    from flask import current_app

    body = request.get_json(force=True) or {}
    records = body.get("records", [])
    if not records:
        return jsonify({"ok": False, "msg": "无数据可导出"})

    try:
        template_path = os.path.join(current_app.static_folder, "天九持股主体公司管理.xlsx")
        if not os.path.exists(template_path):
            return jsonify({"ok": False, "msg": "导出模板文件不存在"})

        _add_log("info", "导出", f"环境={_current_env['value']}, 天九持股主体{len(records)}条")

        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        template_cols = config.HOLDING_SUBJECT_IMPORT_TEMPLATE_COLS
        for row_idx, record in enumerate(records, start=3):
            for col_idx, (_, field, ftype) in enumerate(template_cols, start=1):
                if not field:
                    continue
                val = record.get(field)
                if ftype == "label" and isinstance(val, dict):
                    val = val.get("label", "")
                if val is None:
                    val = ""
                if isinstance(val, (dict, list)):
                    val = json.dumps(val, ensure_ascii=False)
                ws.cell(row=row_idx, column=col_idx, value=val)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        wb.close()

        output = _convert_inline_to_shared_strings(output)

        now_str = datetime.now().strftime("%Y%m%d%H%M")
        download_name = f"导出天九持股主体-{now_str}.xlsx"

        _add_log("info", "导出", f"成功: {download_name}, {len(records)}条数据")
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=download_name,
        )
    except Exception as e:
        _add_log("error", "导出", str(e))
        return jsonify({"ok": False, "msg": f"导出失败：{e}"})


@import_export_bp.route("/api/import/holding-subject/upload", methods=["POST"])
def import_holding_subject_upload():
    """上传天九持股主体导入文件"""
    if _current_env["value"] != "test":
        return jsonify({"ok": False, "msg": "导入功能仅测试环境支持"})

    token = ensure_valid_token()
    if not token:
        return jsonify({"ok": False, "msg": "未登录或 Token 失效"})

    if "file" not in request.files:
        return jsonify({"ok": False, "msg": "未选择文件"})

    file = request.files["file"]
    if not file.filename:
        return jsonify({"ok": False, "msg": "文件名为空"})

    tenant = config.get_tenant(_current_env["value"])
    base_url = config.ENVIRONMENTS[_current_env["value"]]["base_url"]
    upload_url = f"{base_url}/tojoy-form-engine/data/uploadImportExcelFile"

    headers = {"Tj-Auth": token, "Tojoy-Tenant": tenant}
    form_id = "56eb6fb198bf4a4a9882262749409c43"
    files_data = {"files": (file.filename, file.stream,
                            file.content_type or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    form_data = {
        "appId": config.APP_ID,
        "formId": form_id,
        "importMode": "1",
        "saveMode": "1",
    }

    _add_log("info", "导入-上传", f"天九持股主体, 文件={file.filename}")
    try:
        resp = requests.post(upload_url, headers=headers, files=files_data, data=form_data, timeout=60)
        data = resp.json()
        if data.get("code") == 200:
            resp_data = data.get("data") or {}
            batch_no = resp_data.get("batchNo", "") if isinstance(resp_data, dict) else str(resp_data)
            _add_log("info", "导入-上传", f"成功: batchNo={batch_no}")
            return jsonify({"ok": True, "batchNo": batch_no, "data": resp_data, "msg": "文件上传成功"})
        else:
            _add_log("warn", "导入-上传", f"失败: {data.get('msg', '未知错误')}")
            return jsonify({"ok": False, "msg": f"上传失败：{data.get('msg', '未知错误')}"})
    except Exception as e:
        _add_log("error", "导入-上传", str(e))
        return jsonify({"ok": False, "msg": f"上传异常：{e}"})


@import_export_bp.route("/api/import/holding-subject/execute", methods=["POST"])
def import_holding_subject_execute():
    """执行天九持股主体导入"""
    if _current_env["value"] != "test":
        return jsonify({"ok": False, "msg": "导入功能仅测试环境支持"})

    token = ensure_valid_token()
    if not token:
        return jsonify({"ok": False, "msg": "未登录或 Token 失效"})

    body = request.get_json(force=True) or {}
    batch_no = body.get("batchNo", "")
    if not batch_no:
        return jsonify({"ok": False, "msg": "缺少 batchNo"})

    tenant = config.get_tenant(_current_env["value"])
    base_url = config.ENVIRONMENTS[_current_env["value"]]["base_url"]
    form_id = "56eb6fb198bf4a4a9882262749409c43"
    import_url = f"{base_url}/tojoy-form-engine/data/formDataImport"

    headers = {"Content-Type": "application/json", "Tj-Auth": token, "Tojoy-Tenant": tenant}
    payload = {
        "formId": form_id,
        "appId": config.APP_ID,
        "batchNo": batch_no,
        "importMode": "1",
        "saveMode": "1",
    }

    _add_log("info", "导入-执行", f"天九持股主体, batchNo={batch_no}")
    try:
        resp = requests.post(import_url, json=payload, headers=headers, timeout=120)
        data = resp.json()
        resp_data = data.get("data") or {}

        total_num = resp_data.get("totalNum") or resp_data.get("total_num") or 0
        suc_num = resp_data.get("sucNum") or resp_data.get("suc_num") or resp_data.get("checkSucNum") or 0
        fail_num = resp_data.get("failNum") or resp_data.get("fail_num") or resp_data.get("checkFailNum") or 0
        fail_file_url = (resp_data.get("importFailFileUrl") or resp_data.get("failFileUrl")
                         or resp_data.get("fail_file_url") or "")

        is_success = data.get("code") == 200 and fail_num == 0 and suc_num > 0

        if is_success:
            _add_log("info", "导入-执行", f"成功: 共{total_num}行，入库{suc_num}行")
            return jsonify({"ok": True, "data": resp_data, "msg": f"导入成功：共{total_num}行，全部入库"})

        summary = f"共{total_num}行，通过{suc_num}行，失败{fail_num}行"
        if not total_num and not fail_num:
            summary = data.get("msg", "未知错误")

        fail_reasons = []
        if fail_file_url:
            fail_reasons = _parse_fail_file(fail_file_url)

        _add_log("warn", "导入-执行", f"失败: {summary}")
        return jsonify({
            "ok": False,
            "msg": f"导入失败：{summary}",
            "detail": {
                "totalNum": total_num,
                "sucNum": suc_num,
                "failNum": fail_num,
                "failFileUrl": fail_file_url,
                "failReasons": fail_reasons,
            },
        })
    except Exception as e:
        _add_log("error", "导入-执行", str(e))
        return jsonify({"ok": False, "msg": f"导入异常：{e}"})


@import_export_bp.route("/api/export/history-format", methods=["POST"])
def export_history_format():
    """按导入模板格式导出股权历史表数据为 xlsx"""
    from flask import current_app

    body = request.get_json(force=True) or {}
    records = body.get("records", [])
    if not records:
        return jsonify({"ok": False, "msg": "无数据可导出"})

    try:
        template_path = os.path.join(current_app.static_folder,
                                     "股权资产管理历史数据（25年之前的项目股权资产数据）.xlsx")
        if not os.path.exists(template_path):
            return jsonify({"ok": False, "msg": "导出模板文件不存在"})

        _add_log("info", "导出", f"环境={_current_env['value']}, 股权历史表{len(records)}条")

        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        template_cols = config.HISTORY_IMPORT_TEMPLATE_COLS
        for row_idx, record in enumerate(records, start=2):
            for col_idx, (_, field, ftype) in enumerate(template_cols, start=1):
                if not field:
                    continue
                val = record.get(field)
                if ftype == "label" and isinstance(val, dict):
                    val = val.get("label", "")
                if ftype == "date" and val:
                    s = str(val)
                    if s.isdigit() and len(s) >= 10:
                        ts = int(s)
                        if ts > 9999999999:
                            ts = ts / 1000
                        val = datetime.fromtimestamp(ts).strftime("%Y-%m-%d")
                    elif len(s) >= 10 and "-" in s:
                        val = s[:10]
                if val is None:
                    val = ""
                if isinstance(val, (dict, list)):
                    val = json.dumps(val, ensure_ascii=False)
                ws.cell(row=row_idx, column=col_idx, value=val)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        wb.close()

        output = _convert_inline_to_shared_strings(output)

        now_str = datetime.now().strftime("%Y%m%d%H%M")
        download_name = f"导出股权历史-{now_str}.xlsx"

        _add_log("info", "导出", f"成功: {download_name}, {len(records)}条数据")
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=download_name,
        )
    except Exception as e:
        _add_log("error", "导出", str(e))
        return jsonify({"ok": False, "msg": f"导出失败：{e}"})


@import_export_bp.route("/api/import/history/upload", methods=["POST"])
def import_history_upload():
    """上传股权历史表导入文件"""
    if _current_env["value"] != "test":
        return jsonify({"ok": False, "msg": "导入功能仅测试环境支持"})

    token = ensure_valid_token()
    if not token:
        return jsonify({"ok": False, "msg": "未登录或 Token 失效"})

    if "file" not in request.files:
        return jsonify({"ok": False, "msg": "未选择文件"})

    file = request.files["file"]
    if not file.filename:
        return jsonify({"ok": False, "msg": "文件名为空"})

    tenant = config.get_tenant(_current_env["value"])
    base_url = config.ENVIRONMENTS[_current_env["value"]]["base_url"]
    upload_url = f"{base_url}/tojoy-form-engine/data/uploadImportExcelFile"

    headers = {"Tj-Auth": token, "Tojoy-Tenant": tenant}
    form_id = "433011f3e6454bbf9735bb39f643b6c1"
    files_data = {"files": (file.filename, file.stream,
                            file.content_type or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    form_data = {
        "appId": config.APP_ID,
        "formId": form_id,
        "importMode": "1",
        "saveMode": "1",
    }

    _add_log("info", "导入-上传", f"股权历史表, 文件={file.filename}")
    try:
        resp = requests.post(upload_url, headers=headers, files=files_data, data=form_data, timeout=60)
        data = resp.json()
        if data.get("code") == 200:
            resp_data = data.get("data") or {}
            batch_no = resp_data.get("batchNo", "") if isinstance(resp_data, dict) else str(resp_data)
            _add_log("info", "导入-上传", f"成功: batchNo={batch_no}")
            return jsonify({"ok": True, "batchNo": batch_no, "data": resp_data, "msg": "文件上传成功"})
        else:
            _add_log("warn", "导入-上传", f"失败: {data.get('msg', '未知错误')}")
            return jsonify({"ok": False, "msg": f"上传失败：{data.get('msg', '未知错误')}"})
    except Exception as e:
        _add_log("error", "导入-上传", str(e))
        return jsonify({"ok": False, "msg": f"上传异常：{e}"})


@import_export_bp.route("/api/import/history/execute", methods=["POST"])
def import_history_execute():
    """执行股权历史表导入"""
    if _current_env["value"] != "test":
        return jsonify({"ok": False, "msg": "导入功能仅测试环境支持"})

    token = ensure_valid_token()
    if not token:
        return jsonify({"ok": False, "msg": "未登录或 Token 失效"})

    body = request.get_json(force=True) or {}
    batch_no = body.get("batchNo", "")
    if not batch_no:
        return jsonify({"ok": False, "msg": "缺少 batchNo"})

    tenant = config.get_tenant(_current_env["value"])
    base_url = config.ENVIRONMENTS[_current_env["value"]]["base_url"]
    form_id = "433011f3e6454bbf9735bb39f643b6c1"
    import_url = f"{base_url}/tojoy-form-engine/data/formDataImport"

    headers = {"Content-Type": "application/json", "Tj-Auth": token, "Tojoy-Tenant": tenant}
    payload = {
        "formId": form_id,
        "appId": config.APP_ID,
        "batchNo": batch_no,
        "importMode": "1",
        "saveMode": "1",
    }

    _add_log("info", "导入-执行", f"股权历史表, batchNo={batch_no}")
    try:
        resp = requests.post(import_url, json=payload, headers=headers, timeout=120)
        data = resp.json()
        resp_data = data.get("data") or {}

        total_num = resp_data.get("totalNum") or resp_data.get("total_num") or 0
        suc_num = resp_data.get("sucNum") or resp_data.get("suc_num") or resp_data.get("checkSucNum") or 0
        fail_num = resp_data.get("failNum") or resp_data.get("fail_num") or resp_data.get("checkFailNum") or 0
        fail_file_url = (resp_data.get("importFailFileUrl") or resp_data.get("failFileUrl")
                         or resp_data.get("fail_file_url") or "")

        is_success = data.get("code") == 200 and fail_num == 0 and suc_num > 0

        if is_success:
            _add_log("info", "导入-执行", f"成功: 共{total_num}行，入库{suc_num}行")
            return jsonify({"ok": True, "data": resp_data, "msg": f"导入成功：共{total_num}行，全部入库"})

        summary = f"共{total_num}行，通过{suc_num}行，失败{fail_num}行"
        if not total_num and not fail_num:
            summary = data.get("msg", "未知错误")

        fail_reasons = []
        if fail_file_url:
            fail_reasons = _parse_fail_file(fail_file_url)

        _add_log("warn", "导入-执行", f"失败: {summary}")
        return jsonify({
            "ok": False,
            "msg": f"导入失败：{summary}",
            "detail": {
                "totalNum": total_num,
                "sucNum": suc_num,
                "failNum": fail_num,
                "failFileUrl": fail_file_url,
                "failReasons": fail_reasons,
            },
        })
    except Exception as e:
        _add_log("error", "导入-执行", str(e))
        return jsonify({"ok": False, "msg": f"导入异常：{e}"})
