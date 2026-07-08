"""
工具函数模块
提供跨模块使用的通用工具函数。
"""
import io
import re
import zipfile
from datetime import datetime

import openpyxl
import requests

from shared import _add_log


def _convert_inline_to_shared_strings(xlsx_buf):
    """将 openpyxl 生成的 inlineStr 格式 xlsx 转换为 sharedStrings 格式。
    后端 Java POI 导入解析器不支持 inlineStr 类型单元格，
    必须使用 sharedStrings.xml 共享字符串表 + t="s" 引用。
    """
    xlsx_buf.seek(0)
    all_files = {}
    with zipfile.ZipFile(xlsx_buf, 'r') as zin:
        for name in zin.namelist():
            all_files[name] = zin.read(name)

    # 找到 sheet xml
    sheet_path = 'xl/worksheets/sheet1.xml'
    if sheet_path not in all_files:
        xlsx_buf.seek(0)
        return xlsx_buf

    sheet_xml = all_files[sheet_path].decode('utf-8')
    if 't="inlineStr"' not in sheet_xml:
        xlsx_buf.seek(0)
        return xlsx_buf

    # 收集 shared strings
    shared_strings = []
    ss_map = {}

    def get_ss_idx(value):
        if value in ss_map:
            return ss_map[value]
        idx = len(shared_strings)
        shared_strings.append(value)
        ss_map[value] = idx
        return idx

    # 解析已有的 sharedStrings（如果有）
    if 'xl/sharedStrings.xml' in all_files:
        existing_ss = all_files['xl/sharedStrings.xml'].decode('utf-8')
        for m in re.finditer(r'<si><t(?:\s[^>]*)?>([^<]*)</t></si>', existing_ss):
            get_ss_idx(m.group(1))
        if '<si><t/></si>' in existing_ss:
            get_ss_idx('')

    # 替换 inlineStr 为 shared string 引用
    def replace_inline(match):
        attrs = match.group(1)
        body = match.group(2)
        text_match = re.search(r'<is><t(?:\s[^>]*)?>([^<]*)</t></is>', body)
        if text_match:
            value = text_match.group(1)
        else:
            value = ''
        idx = get_ss_idx(value)
        return f'<c {attrs} t="s"><v>{idx}</v></c>'

    new_sheet_xml = re.sub(
        r'<c ([^>]*?)t="inlineStr">(.*?)</c>',
        replace_inline, sheet_xml, flags=re.DOTALL
    )

    # 构建 sharedStrings.xml
    ss_parts = ['<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n']
    ss_parts.append(f'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="{len(shared_strings)}">')
    for s in shared_strings:
        if s == '':
            ss_parts.append('<si><t/></si>')
        elif s.startswith(' ') or s.endswith(' '):
            ss_parts.append(f'<si><t xml:space="preserve">{s}</t></si>')
        else:
            ss_parts.append(f'<si><t>{s}</t></si>')
    ss_parts.append('</sst>')
    shared_strings_xml = ''.join(ss_parts)

    # 确保 Content_Types 包含 sharedStrings
    content_types = all_files['[Content_Types].xml'].decode('utf-8')
    if 'sharedStrings' not in content_types:
        content_types = content_types.replace(
            '</Types>',
            '<Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/></Types>'
        )

    # 确保 workbook.xml.rels 包含 sharedStrings 关系
    wb_rels_path = 'xl/_rels/workbook.xml.rels'
    wb_rels = all_files[wb_rels_path].decode('utf-8')
    if 'sharedStrings' not in wb_rels:
        rid_nums = [int(x) for x in re.findall(r'rId(\d+)', wb_rels)]
        next_rid = max(rid_nums) + 1 if rid_nums else 3
        wb_rels = wb_rels.replace(
            '</Relationships>',
            f'<Relationship Id="rId{next_rid}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" Target="sharedStrings.xml"/></Relationships>'
        )

    # 重新打包 xlsx
    result = io.BytesIO()
    with zipfile.ZipFile(result, 'w', zipfile.ZIP_DEFLATED) as zout:
        for name, data in all_files.items():
            if name == sheet_path:
                zout.writestr(name, new_sheet_xml.encode('utf-8'))
            elif name == '[Content_Types].xml':
                zout.writestr(name, content_types.encode('utf-8'))
            elif name == wb_rels_path:
                zout.writestr(name, wb_rels.encode('utf-8'))
            elif name == 'xl/sharedStrings.xml':
                continue  # 用新生成的替换
            else:
                zout.writestr(name, data)
        zout.writestr('xl/sharedStrings.xml', shared_strings_xml.encode('utf-8'))

    result.seek(0)
    return result


def _parse_fail_file(file_url):
    """下载并解析导入失败详情文件（xlsx），提取每行的失败原因"""
    reasons = []
    try:
        resp = requests.get(file_url, timeout=30)
        if resp.status_code != 200:
            return reasons
        wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True)
        ws = wb.active
        if not ws:
            return reasons

        # 找到表头行，定位"失败原因"列
        header_row = None
        reason_col = None
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row and any("失败原因" in str(cell or "") for cell in row):
                header_row = row_idx
                for col_idx, cell in enumerate(row):
                    if cell and "失败原因" in str(cell):
                        reason_col = col_idx
                        break
                break

        if reason_col is None:
            # 没找到"失败原因"列，尝试取最后一列
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row:
                    last_val = str(row[-1] or "").strip()
                    if last_val:
                        reasons.append(last_val)
        else:
            # 从数据行开始读取失败原因
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                if row and len(row) > reason_col:
                    val = str(row[reason_col] or "").strip()
                    if val:
                        reasons.append(val)

        wb.close()
    except Exception as e:
        _add_log("error", "导入-执行", f"解析失败文件异常: {e}")
    return reasons


def _parse_date_to_str(value):
    """将日期值（时间戳或字符串）转为 YYYY-MM-DD 格式字符串"""
    if value is None:
        return None
    s = str(value)
    # 时间戳（毫秒或秒）
    if s.isdigit() and len(s) >= 10:
        ts = int(s)
        if ts > 9999999999:
            ts = ts / 1000
        d = datetime.fromtimestamp(ts)
        return d.strftime("%Y-%m-%d")
    # 已经是日期字符串
    if len(s) >= 10 and "-" in s:
        return s[:10]
    return None


def _safe_number(value):
    """安全提取数值，处理 None/空字符串/非数字情况"""
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None
